"""エンドツーエンドの自動テスト（実データ不要）。

ダミーのPDF(損保ジャパン様式を模擬)とダミーのExcelをその場で生成し、
「PDF解析 → 拠点照合 → 入力シートへ追記」までを通しで検証する。
実際の保険データが無くても、システム全体の動作を確認できる。
"""
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl  # noqa: E402

from core import excel_writer, pipeline  # noqa: E402

# ダミー明細 (証券番号, 件数, 保険料, 入金日, 累計, 営業所コード)
DUMMY = [
    ("12AB00001", "1", "24,190", "R8.07.16", "24,190", "038"),
    ("12AB00002", "1", "17,650", "R8.07.16", "41,840", "038"),
    ("12CD00003", "1", "18,160", "R8.07.17", "60,000", "030"),
    # 書損(保険料・日付なし)= 除外されるべき行
    ("12ZZ99999", "", "", "", "", "030"),
]
# 列 x 座標 (実PDFに合わせた位置)
COLX = {"cert": 35, "ken": 141, "prem": 225, "date": 283, "cum": 419, "code": 479}


def _make_pdf(path: Path) -> None:
    """ダミー明細を配置した最小PDFを生成（ASCIIのみ）。"""
    def text(x, y, s):
        return f"BT /F1 9 Tf {x} {y} Td ({s}) Tj ET\n"

    content = text(40, 770, "TESTSJ jibaiseki meisai marker")
    y = 720
    for cert, ken, prem, date, cum, code in DUMMY:
        for key, val in [("cert", cert), ("ken", ken), ("prem", prem),
                         ("date", date), ("cum", cum), ("code", code)]:
            if val:
                content += text(COLX[key], y, val)
        y -= 18

    stream = content.encode("latin-1")
    objs = []
    objs.append(b"<</Type/Catalog/Pages 2 0 R>>")
    objs.append(b"<</Type/Pages/Kids[3 0 R]/Count 1>>")
    objs.append(b"<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]"
                b"/Contents 4 0 R/Resources<</Font<</F1 5 0 R>>>>>>")
    objs.append(b"<</Length %d>>stream\n" % len(stream) + stream + b"\nendstream ")
    objs.append(b"<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>")

    pdf = b"%PDF-1.4\n"
    offsets = []
    for i, o in enumerate(objs, 1):
        offsets.append(len(pdf))
        pdf += b"%d 0 obj" % i + o + b"endobj\n"
    xref = len(pdf)
    pdf += b"xref\n0 %d\n" % (len(objs) + 1)
    pdf += b"0000000000 65535 f \n"
    for off in offsets:
        pdf += b"%010d 00000 n \n" % off
    pdf += b"trailer<</Root 1 0 R/Size %d>>\nstartxref\n%d\n%%%%EOF" % (len(objs) + 1, xref)
    path.write_bytes(pdf)


def _make_workbook(path: Path) -> None:
    """入力マスタと入力シートを持つダミーExcelを生成。"""
    wb = openpyxl.Workbook()
    # 入力マスタ (F=コード, G=事業部, H=拠点, 3行目から)
    m = wb.active
    m.title = "入力マスタ"
    m["F2"] = "拠点コード"; m["G2"] = "事業部"; m["H2"] = "拠点"
    for r, (code, div, base) in enumerate(
        [("030", "ボルボ", "千里"), ("038", "ボルボ", "板橋")], start=3
    ):
        m[f"F{r}"] = code; m[f"G{r}"] = div; m[f"H{r}"] = base
    # 入力シート (見出し5行目, データ6行目〜)
    ws = wb.create_sheet("自賠責（入力シート）")
    hdr = ["月", "日", "月日", "拠点コード", "事業部", "拠点", "保険会社", "保険料", "手数料", "正味保険料"]
    for c, h in enumerate(hdr, 1):
        ws.cell(row=5, column=c, value=h)
    ws.cell(row=6, column=1, value="既存")  # 既存データ1行
    wb.save(path)


def _test_config():
    companies = {"テスト損保": {
        "detect_keyword": "TESTSJ", "company_name": "損保ジャパン", "fee_per_unit": 1735,
        "columns": {"cert": 120, "ken": 200, "prem": 270, "date": 360, "cum": 460, "code": 99999},
    }}
    settings = {
        "target_sheet": "自賠責（入力シート）", "header_row": 5,
        "columns": {"月": "A", "日": "B", "月日": "C", "拠点コード": "D", "事業部": "E",
                    "拠点": "F", "保険会社": "G", "保険料": "H", "手数料": "I", "正味保険料": "J"},
        "master_sheet": "入力マスタ", "master_code_col": "F", "master_div_col": "G",
        "master_base_col": "H", "master_start_row": 3,
    }
    return companies, settings


def test_end_to_end():
    tmp = Path(tempfile.mkdtemp())
    pdf_dir = tmp / "pdf"; pdf_dir.mkdir()
    _make_pdf(pdf_dir / "test.pdf")
    wb_path = tmp / "台帳.xlsx"
    _make_workbook(wb_path)

    companies, settings = _test_config()
    comp_yaml = tmp / "companies.yaml"
    import yaml
    comp_yaml.write_text(yaml.safe_dump(companies, allow_unicode=True), encoding="utf-8")

    rows, reports = pipeline.process_dir(pdf_dir, wb_path, comp_yaml, settings)

    # 書損行は除外され、3件のはず
    assert len(rows) == 3, f"件数が想定外: {len(rows)}"
    # 合計 = 24190 + 17650 + 18160
    assert sum(r["保険料"] for r in rows) == 60000
    # 拠点照合
    assert rows[0]["拠点"] == "板橋" and rows[0]["事業部"] == "ボルボ"
    assert rows[2]["拠点"] == "千里"
    # 日付・計算
    assert rows[0]["月日"] == "7月16日"
    assert rows[0]["正味保険料"] == 24190 - 1735

    # Excelへ追記
    out = tmp / "out.xlsx"
    saved, start, n = excel_writer.append_rows(wb_path, rows, settings, out)
    assert n == 3 and start == 7  # 既存6行目の次=7行目から

    ws = openpyxl.load_workbook(saved, data_only=True)["自賠責（入力シート）"]
    assert ws.cell(row=7, column=6).value == "板橋"
    assert ws.cell(row=7, column=10).value == 22455
    assert ws.cell(row=9, column=6).value == "千里"


if __name__ == "__main__":
    test_end_to_end()
    print("✓ エンドツーエンドテスト成功（PDF生成→解析→照合→Excel追記まで通し）")
