"""抽出結果を集計 Excel に追記するモジュール。"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import re

import openpyxl
from openpyxl.utils import column_index_from_string

# 列文字指定(A, B, .. ZZ)かどうかの判定用（ASCII 英字のみ）
_COL_LETTER_RE = re.compile(r"^[A-Za-z]{1,3}$")


def load_column_config(path: str | Path) -> dict[str, Any]:
    import yaml
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _resolve_columns(ws, cfg: dict[str, Any]) -> dict[str, int]:
    """抽出項目名 -> Excel の列番号 の対応を作る。

    columns の値が "A" のような列文字ならそのまま列指定として扱い、
    それ以外は header_row からその見出しを持つ列を探す。
    """
    header_row = cfg.get("header_row", 1)
    mapping: dict[str, int] = {}

    # 見出し行を読み取っておく
    headers: dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.column

    for field, target in (cfg.get("columns") or {}).items():
        target = str(target).strip()
        # 列文字指定 (A, B, ...) か? (ASCII 英字のみを列指定とみなす)
        if _COL_LETTER_RE.match(target):
            mapping[field] = column_index_from_string(target.upper())
        elif target in headers:
            mapping[field] = headers[target]
        # 見つからない項目はスキップ(書き込まない)
    return mapping, headers


def _next_row(ws, cfg: dict[str, Any]) -> int:
    if cfg.get("start_row"):
        return int(cfg["start_row"])
    # 既存データの最終行の次
    header_row = cfg.get("header_row", 1)
    last = header_row
    for row in range(header_row + 1, ws.max_row + 1):
        if any(ws.cell(row=row, column=c).value not in (None, "") for c in range(1, ws.max_column + 1)):
            last = row
    return last + 1


def write_rows(
    excel_path: str | Path,
    records: list[dict[str, str]],
    cfg: dict[str, Any],
    output_path: str | Path | None = None,
) -> Path:
    """records を Excel に追記して保存する。

    excel_path   : 既存の集計 Excel
    records      : extract() が返す辞書のリスト
    cfg          : columns.yaml の内容
    output_path  : 保存先。None なら excel_path を上書き。
    戻り値: 保存したファイルパス
    """
    excel_path = Path(excel_path)
    wb = openpyxl.load_workbook(excel_path)

    sheet_name = cfg.get("sheet") or ""
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    mapping, headers = _resolve_columns(ws, cfg)
    add_source = cfg.get("add_source", False)
    source_col = None
    if add_source:
        source_col = headers.get("元PDF")
        if source_col is None:
            # 見出しが無ければ最終列の次に追加
            source_col = (max(headers.values()) + 1) if headers else 1
            ws.cell(row=cfg.get("header_row", 1), column=source_col, value="元PDF")

    row = _next_row(ws, cfg)
    for rec in records:
        for field, col in mapping.items():
            val = rec.get(field, "")
            if val != "":
                ws.cell(row=row, column=col, value=val)
        if source_col is not None and rec.get("_source"):
            ws.cell(row=row, column=source_col, value=rec["_source"])
        row += 1

    out = Path(output_path) if output_path else excel_path
    wb.save(out)
    return out
